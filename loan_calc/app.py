# Импорт необходимых библиотек
from flask import Flask, render_template, request, send_file  # Flask для веб-приложения
import math  # Математические функции
from decimal import Decimal, getcontext  # Точные вычисления с плавающей точкой
from openpyxl import Workbook  # Создание Excel файлов
from openpyxl.styles import Font, Alignment, PatternFill  # Стили для Excel
import io  # Работа с байтовыми потоками
from datetime import datetime  # Работа с датами

# Создание экземпляра Flask приложения
app = Flask(__name__)

# Устанавливаем точность для расчетов с плавающей точкой (28 знаков)
# Это необходимо для точных финансовых вычислений
getcontext().prec = 28

def format_number(n):
    """
    Функция для форматирования чисел с пробелами для разделения разрядов
    Например: 1000000 -> 1 000 000
    
    Args:
        n: Число для форматирования
        
    Returns:
        str: Отформатированное число с пробелами
    """
    if isinstance(n, float) and not n.is_integer():
        # Для дробных чисел сохраняем 2 знака после запятой
        s = f"{n:,.2f}"
    else:
        # Для целых чисел убираем .00
        s = f"{int(n):,}"
    # Заменяем запятые на пробелы и убираем .00
    return s.replace(",", " ").replace(".00", "")

def create_excel_file(schedule, loan_data, early_repayment_data=None):
    """
    Создает Excel файл с графиком платежей и дополнительной информацией
    
    Args:
        schedule: График платежей
        loan_data: Данные кредита (сумма, срок, ставка и т.д.)
        early_repayment_data: Данные досрочного погашения (опционально)
        
    Returns:
        BytesIO: Файл в байтовом формате для скачивания
    """
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active  # Получаем активный лист
    ws.title = "График платежей"  # Устанавливаем название листа
    
    # Создаем заголовок документа
    title_cell = ws.cell(row=1, column=1, value="Ипотечный калькулятор")
    title_cell.font = Font(size=20, bold=True)  # Жирный шрифт 20pt
    title_cell.alignment = Alignment(horizontal='center')  # Центрирование
    
    # Рассчитываем параметры кредита для отображения
    row = 3  # Начинаем с 3-й строки
    loan_amount = loan_data['principal'] - loan_data['down_payment']  # Сумма кредита
    months = loan_data['years'] * 12  # Общее количество месяцев
    monthly_rate = Decimal(str(loan_data['rate'])) / Decimal('100') / Decimal('12')  # Месячная ставка
    
    # Рассчитываем ежемесячный платеж по формуле аннуитета
    if monthly_rate == 0:
        # Если ставка 0%, то просто делим сумму на количество месяцев
        monthly_payment = loan_amount / months
    else:
        # Формула аннуитетного платежа: PMT = PV * r * (1 + r)^n / ((1 + r)^n - 1)
        numerator = monthly_rate * (Decimal('1') + monthly_rate) ** months
        denominator = (Decimal('1') + monthly_rate) ** months - Decimal('1')
        monthly_payment = float(Decimal(str(loan_amount)) * numerator / denominator)
    
    # Рассчитываем общие суммы
    total_payment = monthly_payment * months  # Общая сумма выплат
    total_interest = total_payment - loan_amount  # Переплата по процентам
    
    # Создаем список параметров для отображения в Excel
    params = [
        ("Стоимость недвижимости:", f"{loan_data['principal']:,.2f} ₽"),
        ("Первоначальный взнос:", f"{loan_data['down_payment']:,.2f} ₽"),
        ("Сумма кредита:", f"{loan_amount:,.2f} ₽"),
        ("Срок кредита:", f"{loan_data['years']} лет"),
        ("Процентная ставка:", f"{loan_data['rate']}% годовых"),
        ("Ежемесячный платеж:", f"{schedule[0]['payment']:,.2f} ₽"),
        ("Общая сумма выплат:", f"{total_payment:,.2f} ₽"),
        ("Переплата по процентам:", f"{total_interest:,.2f} ₽")
    ]
    
    # Записываем параметры в Excel
    for param_name, param_value in params:
        ws.cell(row=row, column=1, value=param_name).font = Font(bold=True)
        ws.cell(row=row, column=2, value=param_value)
        row += 1
    
    # Добавляем пустую строку перед таблицей
    row += 1
    
    # Создаем заголовки таблицы
    headers = ['Год', 'Месяц', 'Ежемесячный платеж', 'Основной долг', 'Проценты', 'Остаток долга']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)  # Жирный шрифт
        cell.alignment = Alignment(horizontal='center')  # Центрирование
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")  # Серый фон
    
    # Записываем данные графика платежей
    for payment in schedule:
        row += 1
        ws.cell(row=row, column=1, value=payment['year'])
        ws.cell(row=row, column=2, value=payment['month'])
        ws.cell(row=row, column=3, value=round(payment['payment'], 2))
        ws.cell(row=row, column=4, value=round(payment['principal'], 2))
        ws.cell(row=row, column=5, value=round(payment['interest'], 2))
        ws.cell(row=row, column=6, value=round(payment['remaining_balance'], 2))
    
    # Автоматически подбираем ширину столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)  # Ограничиваем максимальную ширину
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Объединяем ячейки для заголовка (A1:F1)
    ws.merge_cells('A1:F1')
    
    # Добавляем лист досрочного погашения, если есть данные
    if early_repayment_data:
        ws_early = wb.create_sheet("Досрочное погашение")
        
        # Заголовок листа досрочного погашения
        title_cell = ws_early.cell(row=1, column=1, value="Досрочное погашение")
        title_cell.font = Font(size=20, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Параметры досрочного погашения
        row = 3
        early_params = [
            ("Сумма досрочного погашения:", f"{early_repayment_data['amount']:,.2f} ₽"),
            ("Дата досрочного погашения:", f"{early_repayment_data['month']}/{early_repayment_data['year']}"),
            ("Режим:", "Уменьшить платеж" if early_repayment_data['mode'] == 'reduce_payment' else "Сократить срок"),
            ("", ""),
            ("Сравнение результатов:", ""),
            ("Исходный ежемесячный платеж:", f"{early_repayment_data['original_payment']:,.2f} ₽"),
            ("Новый ежемесячный платеж:", f"{early_repayment_data['new_payment']:,.2f} ₽"),
            ("Экономия на процентах:", f"{early_repayment_data['interest_savings']:,.2f} ₽"),
            ("Сокращение срока:", f"{early_repayment_data['term_reduction']:.1f} лет")
        ]
        
        # Записываем параметры досрочного погашения
        for param_name, param_value in early_params:
            ws_early.cell(row=row, column=1, value=param_name).font = Font(bold=True)
            ws_early.cell(row=row, column=2, value=param_value)
            row += 1
        
        # Создаем сравнительную таблицу
        row += 2
        headers = ['Год', 'Исходный платеж', 'Новый платеж', 'Разница']
        for col, header in enumerate(headers, 1):
            cell = ws_early.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Записываем данные сравнения
        for i, (orig, new) in enumerate(zip(early_repayment_data['original_schedule'], early_repayment_data['new_schedule'])):
            row += 1
            ws_early.cell(row=row, column=1, value=orig['year'])
            ws_early.cell(row=row, column=2, value=round(orig['payment'], 2))
            ws_early.cell(row=row, column=3, value=round(new['payment'], 2))
            ws_early.cell(row=row, column=4, value=round(orig['payment'] - new['payment'], 2))
        
        # Автоматически подбираем ширину столбцов для листа досрочного погашения
        for column in ws_early.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws_early.column_dimensions[column_letter].width = adjusted_width
        
        # Объединяем ячейки для заголовка
        ws_early.merge_cells('A1:D1')
    
    # Сохраняем в байтовый поток для отправки пользователю
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  # Перемещаем указатель в начало файла
    return excel_file

def calculate_payment_schedule(principal, down_payment, years, rate):
    """
    Рассчитывает график платежей по кредиту
    
    Args:
        principal: Стоимость недвижимости
        down_payment: Первоначальный взнос
        years: Срок кредита в годах
        rate: Процентная ставка годовых
        
    Returns:
        tuple: (график платежей, итоговые суммы)
    """
    # Рассчитываем основные параметры
    loan_amount = principal - down_payment  # Сумма кредита
    months = years * 12  # Общее количество месяцев
    monthly_rate = Decimal(str(rate)) / Decimal('100') / Decimal('12')  # Месячная ставка
    
    # Рассчитываем ежемесячный платеж
    if monthly_rate == 0:
        # Если ставка 0%, то просто делим сумму на количество месяцев
        monthly_payment = loan_amount / months
    else:
        # Формула аннуитетного платежа
        numerator = monthly_rate * (Decimal('1') + monthly_rate) ** months
        denominator = (Decimal('1') + monthly_rate) ** months - Decimal('1')
        monthly_payment = float(Decimal(str(loan_amount)) * numerator / denominator)
    
    # Инициализируем переменные для расчета графика
    schedule = []  # График платежей
    remaining_balance = loan_amount  # Остаток долга
    total_interest_paid = 0  # Общая сумма выплаченных процентов
    total_principal_paid = 0  # Общая сумма выплаченного основного долга
    
    # Рассчитываем каждый месяц
    for month in range(1, months + 1):
        # Рассчитываем проценты за текущий месяц
        interest_payment = float(Decimal(str(remaining_balance)) * monthly_rate)
        
        # Рассчитываем выплату основного долга
        principal_payment = monthly_payment - interest_payment
        
        # Корректировка для последнего платежа (чтобы точно погасить долг)
        if month == months:
            principal_payment = remaining_balance
            monthly_payment = principal_payment + interest_payment
        
        # Обновляем остаток долга и накопительные суммы
        remaining_balance -= principal_payment
        total_interest_paid += interest_payment
        total_principal_paid += principal_payment
        
        # Добавляем только каждый 12-й месяц (годовые данные) для таблицы
        # Также добавляем первый месяц
        if month % 12 == 0 or month == 1:
            schedule.append({
                'month': month,
                'year': (month - 1) // 12 + 1,  # Рассчитываем год
                'payment': monthly_payment,
                'principal': principal_payment,
                'interest': interest_payment,
                'remaining_balance': max(0, remaining_balance),  # Не может быть отрицательным
                'total_interest_paid': total_interest_paid,
                'total_principal_paid': total_principal_paid
            })
    
    return schedule, {
        'total_interest': total_interest_paid,
        'total_principal': total_principal_paid,
        'total_payment': total_interest_paid + total_principal_paid
    }

def calculate_overpayment(principal, down_payment, years, rate):
    """
    Рассчитывает переплату по ипотеке и другие параметры
    
    Args:
        principal: Стоимость недвижимости
        down_payment: Первоначальный взнос
        years: Срок кредита в годах
        rate: Процентная ставка годовых
        
    Returns:
        dict: Словарь с результатами расчета
    """
    # Рассчитываем сумму кредита после первоначального взноса
    loan_amount = principal - down_payment
    months = years * 12
    monthly_rate = Decimal(str(rate)) / Decimal('100') / Decimal('12')
    
    # Рассчитываем ежемесячный платеж
    if monthly_rate == 0:
        monthly_payment = loan_amount / months
    else:
        # Правильная формула аннуитетного платежа
        numerator = monthly_rate * (Decimal('1') + monthly_rate) ** months
        denominator = (Decimal('1') + monthly_rate) ** months - Decimal('1')
        monthly_payment = float(Decimal(str(loan_amount)) * numerator / denominator)
    
    # Рассчитываем общие суммы
    total_payment = monthly_payment * months  # Общая сумма выплат
    overpayment = total_payment - loan_amount  # Переплата по процентам
    required_income = monthly_payment * 2.5  # Примерный необходимый доход (платеж * 2.5)
    
    # Получаем график платежей
    schedule, totals = calculate_payment_schedule(principal, down_payment, years, rate)
    
    return {
        'overpayment': overpayment,
        'monthly_payment': monthly_payment,
        'total_payment': total_payment,
        'required_income': required_income,
        'loan_amount': loan_amount,
        'schedule': schedule,
        'totals': totals
    }

def calculate_early_repayment(principal, down_payment, years, rate, early_amount, early_month, early_year, mode):
    """
    Рассчитывает параметры досрочного погашения
    
    Args:
        principal: Стоимость недвижимости
        down_payment: Первоначальный взнос
        years: Срок кредита в годах
        rate: Процентная ставка годовых
        early_amount: Сумма досрочного погашения
        early_month: Месяц досрочного погашения
        early_year: Год досрочного погашения
        mode: Режим досрочного погашения ('reduce_payment' или 'reduce_term')
        
    Returns:
        tuple: (результат расчета, ошибка)
    """
    loan_amount = principal - down_payment
    monthly_rate = Decimal(str(rate)) / Decimal('100') / Decimal('12')
    
    # Рассчитываем исходный ежемесячный платеж
    if monthly_rate == 0:
        original_monthly_payment = loan_amount / (years * 12)
    else:
        numerator = monthly_rate * (Decimal('1') + monthly_rate) ** (years * 12)
        denominator = (Decimal('1') + monthly_rate) ** (years * 12) - Decimal('1')
        original_monthly_payment = float(Decimal(str(loan_amount)) * numerator / denominator)
    
    # Определяем месяц досрочного погашения (относительно начала кредита)
    current_year = datetime.now().year
    months_until_early = (early_year - current_year) * 12 + (early_month - 1)
    
    # Проверяем, что дата не в прошлом
    if months_until_early < 0:
        return None, "Дата досрочного погашения не может быть в прошлом"
    
    # Рассчитываем остаток долга на момент досрочного погашения
    remaining_balance_at_early = loan_amount
    for month in range(1, months_until_early + 1):
        interest_payment = float(Decimal(str(remaining_balance_at_early)) * monthly_rate)
        principal_payment = original_monthly_payment - interest_payment
        remaining_balance_at_early -= principal_payment
    
    # Применяем досрочное погашение
    new_balance = remaining_balance_at_early - early_amount
    
    # Проверяем, что сумма досрочного погашения не превышает остаток долга
    if new_balance <= 0:
        return None, "Сумма досрочного погашения слишком велика"
    
    remaining_months = (years * 12) - months_until_early
    
    if mode == "reduce_payment":
        # Режим: уменьшаем платеж, срок остается прежним
        if monthly_rate == 0:
            new_monthly_payment = new_balance / remaining_months
        else:
            numerator = monthly_rate * (Decimal('1') + monthly_rate) ** remaining_months
            denominator = (Decimal('1') + monthly_rate) ** remaining_months - Decimal('1')
            new_monthly_payment = float(Decimal(str(new_balance)) * numerator / denominator)
        
        new_term = years
        term_reduction = 0
        
    else:  # reduce_term
        # Режим: сокращаем срок, платеж остается прежним
        if monthly_rate == 0:
            new_term_months = new_balance / original_monthly_payment
        else:
            # Решаем уравнение для нахождения нового количества месяцев
            # new_balance = original_monthly_payment * (1 - (1 + r)^(-n)) / r
            # где r - месячная ставка, n - количество месяцев
            if monthly_rate == 0:
                new_term_months = new_balance / original_monthly_payment
            else:
                # Используем приближенную формулу для нахождения n
                monthly_rate_float = float(monthly_rate)
                new_term_months = -math.log(1 - (new_balance * monthly_rate_float / original_monthly_payment)) / math.log(1 + monthly_rate_float)
        
        new_term_months = max(1, int(new_term_months))
        new_monthly_payment = original_monthly_payment
        new_term = months_until_early / 12 + new_term_months / 12
        term_reduction = years - new_term
    
    # Рассчитываем экономию на процентах
    # Получаем исходный график платежей
    original_schedule, original_totals = calculate_payment_schedule(principal, down_payment, years, rate)
    
    # Новый график платежей после досрочного погашения
    new_schedule, new_totals = calculate_payment_schedule_after_early(
        principal, down_payment, years, rate, early_amount, early_month, early_year, mode
    )
    
    # Рассчитываем проценты, которые уже выплачены до досрочного погашения
    interest_paid_before_early = 0
    for payment in original_schedule:
        if payment['month'] <= months_until_early:
            interest_paid_before_early += payment['interest']
    
    # Экономия = (общие проценты по исходному графику - проценты до досрочного погашения) - новые проценты
    interest_savings = (original_totals['total_interest'] - interest_paid_before_early) - new_totals['total_interest']
    
    print(f"DEBUG: original_totals['total_interest']: {original_totals['total_interest']}")
    print(f"DEBUG: interest_paid_before_early: {interest_paid_before_early}")
    print(f"DEBUG: new_totals['total_interest']: {new_totals['total_interest']}")
    print(f"DEBUG: interest_savings: {interest_savings}")
    
    return {
        'new_monthly_payment': new_monthly_payment,
        'new_term': new_term,
        'interest_savings': interest_savings,
        'term_reduction': term_reduction,
        'new_schedule': new_schedule
    }, None

def calculate_payment_schedule_after_early(principal, down_payment, years, rate, early_amount, early_month, early_year, mode):
    """
    Рассчитывает график платежей после досрочного погашения
    
    Args:
        principal: Стоимость недвижимости
        down_payment: Первоначальный взнос
        years: Срок кредита в годах
        rate: Процентная ставка годовых
        early_amount: Сумма досрочного погашения
        early_month: Месяц досрочного погашения
        early_year: Год досрочного погашения
        mode: Режим досрочного погашения
        
    Returns:
        tuple: (график платежей, итоговые суммы)
    """
    loan_amount = principal - down_payment
    months = years * 12
    monthly_rate = Decimal(str(rate)) / Decimal('100') / Decimal('12')
    
    # Рассчитываем исходный ежемесячный платеж
    if monthly_rate == 0:
        original_monthly_payment = loan_amount / months
    else:
        numerator = monthly_rate * (Decimal('1') + monthly_rate) ** months
        denominator = (Decimal('1') + monthly_rate) ** months - Decimal('1')
        original_monthly_payment = float(Decimal(str(loan_amount)) * numerator / denominator)
    
    # Определяем месяц досрочного погашения
    current_year = datetime.now().year
    months_until_early = (early_year - current_year) * 12 + (early_month - 1)
    
    # Инициализируем переменные для расчета нового графика
    schedule = []
    remaining_balance = loan_amount
    total_interest_paid = 0
    total_principal_paid = 0
    
    # Рассчитываем каждый месяц с учетом досрочного погашения
    for month in range(1, months + 1):
        # Рассчитываем проценты за текущий месяц
        interest_payment = float(Decimal(str(remaining_balance)) * monthly_rate)
        
        # Применяем досрочное погашение в нужный месяц
        if month == months_until_early + 1:
            remaining_balance -= early_amount
            if remaining_balance <= 0:
                break
        
        # Рассчитываем ежемесячный платеж в зависимости от режима
        if mode == "reduce_payment":
            # После досрочного погашения пересчитываем платеж
            if month > months_until_early:
                remaining_months = months - month + 1
                if monthly_rate == 0:
                    monthly_payment = remaining_balance / remaining_months
                else:
                    numerator = monthly_rate * (Decimal('1') + monthly_rate) ** remaining_months
                    denominator = (Decimal('1') + monthly_rate) ** remaining_months - Decimal('1')
                    monthly_payment = float(Decimal(str(remaining_balance)) * numerator / denominator)
            else:
                monthly_payment = original_monthly_payment
        else:
            # Режим сокращения срока - платеж остается прежним
            monthly_payment = original_monthly_payment
        
        # Рассчитываем выплату основного долга
        principal_payment = monthly_payment - interest_payment
        
        # Корректировка для последнего платежа
        if month == months:
            principal_payment = remaining_balance
            monthly_payment = principal_payment + interest_payment
        
        # Обновляем остаток долга и накопительные суммы
        remaining_balance -= principal_payment
        total_interest_paid += interest_payment
        total_principal_paid += principal_payment
        
        # Добавляем данные в график (каждый 12-й месяц, первый месяц и месяц досрочного погашения)
        if month % 12 == 0 or month == 1 or month == months_until_early + 1:
            schedule.append({
                'month': month,
                'year': (month - 1) // 12 + 1,
                'payment': monthly_payment,
                'principal': principal_payment,
                'interest': interest_payment,
                'remaining_balance': max(0, remaining_balance),
                'total_interest_paid': total_interest_paid,
                'total_principal_paid': total_principal_paid,
                'is_early_repayment': month == months_until_early + 1
            })
        
        # Если долг полностью погашен, прекращаем расчет
        if remaining_balance <= 0:
            break
    
    return schedule, {
        'total_interest': total_interest_paid,
        'total_principal': total_principal_paid,
        'total_payment': total_interest_paid + total_principal_paid
    }

@app.route('/', methods=['GET', 'POST'])
def index():
    """
    Главная страница приложения
    Обрабатывает GET и POST запросы
    """
    result = None  # Сообщение об ошибке
    values = {'principal': '', 'down_payment': '', 'years': '', 'rate': ''}  # Значения полей формы
    data = None  # Результаты расчета
    
    if request.method == 'POST':
        try:
            # Получаем данные из формы и очищаем от пробелов
            principal = float(request.form['principal'].replace(' ',''))
            down_payment = float(request.form['down_payment'].replace(' ',''))
            years = int(request.form['years'].replace(' ',''))
            rate = float(request.form['rate'].replace(' ',''))
            
            # Форматируем значения для отображения в форме
            values = {
                'principal': format_number(principal),
                'down_payment': format_number(down_payment),
                'years': years,
                'rate': rate
            }
            
            # Выполняем расчет
            data = calculate_overpayment(principal, down_payment, years, rate)
            
        except Exception as e:
            # В случае ошибки сохраняем сообщение
            result = f"Ошибка: {e}"
    
    # Отображаем страницу с результатами
    return render_template('index.html', result=result, data=data, values=values, format_number=format_number)

@app.route('/download_excel', methods=['POST'])
def download_excel():
    """
    Обрабатывает запрос на скачивание Excel файла
    """
    try:
        # Получаем данные из формы
        principal = float(request.form['principal'].replace(' ',''))
        down_payment = float(request.form['down_payment'].replace(' ',''))
        years = int(request.form['years'].replace(' ',''))
        rate = float(request.form['rate'].replace(' ',''))
        
        # Рассчитываем график платежей
        schedule, totals = calculate_payment_schedule(principal, down_payment, years, rate)
        
        # Проверяем, есть ли данные досрочного погашения
        early_repayment_data = None
        if 'early_amount' in request.form and request.form['early_amount'].strip():
            early_amount = float(request.form['early_amount'].replace(' ',''))
            early_month = int(request.form['early_month'])
            early_year = int(request.form['early_year'])
            mode = request.form.get('repayment_mode', 'reduce_payment')
            
            # Рассчитываем досрочное погашение
            result, error = calculate_early_repayment(
                principal, down_payment, years, rate, 
                early_amount, early_month, early_year, mode
            )
            
            if not error:
                # Получаем исходный график
                original_schedule, _ = calculate_payment_schedule(principal, down_payment, years, rate)
                
                # Рассчитываем исходный ежемесячный платеж
                loan_amount = principal - down_payment
                months = years * 12
                monthly_rate = Decimal(str(rate)) / Decimal('100') / Decimal('12')
                
                if monthly_rate == 0:
                    original_monthly_payment = loan_amount / months
                else:
                    numerator = monthly_rate * (Decimal('1') + monthly_rate) ** months
                    denominator = (Decimal('1') + monthly_rate) ** months - Decimal('1')
                    original_monthly_payment = float(Decimal(str(loan_amount)) * numerator / denominator)
                
                # Формируем данные для Excel
                early_repayment_data = {
                    'amount': early_amount,
                    'month': early_month,
                    'year': early_year,
                    'mode': mode,
                    'original_payment': original_monthly_payment,
                    'new_payment': result['new_monthly_payment'],
                    'interest_savings': result['interest_savings'],
                    'term_reduction': result['term_reduction'],
                    'original_schedule': original_schedule,
                    'new_schedule': result['new_schedule']
                }
        
        # Создаем Excel файл
        excel_file = create_excel_file(schedule, {
            'principal': principal,
            'down_payment': down_payment,
            'years': years,
            'rate': rate
        }, early_repayment_data)
        
        # Генерируем имя файла с текущей датой и временем
        filename = f"график_платежей_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Отправляем файл пользователю
        return send_file(
            excel_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return f"Ошибка при создании файла: {e}", 400

@app.route('/calculate_early_repayment', methods=['POST'])
def calculate_early_repayment_route():
    """
    API endpoint для расчета досрочного погашения
    Возвращает JSON с результатами
    """
    try:
        # Получаем данные из формы
        principal = float(request.form['principal'].replace(' ',''))
        down_payment = float(request.form['down_payment'].replace(' ',''))
        years = int(request.form['years'].replace(' ',''))
        rate = float(request.form['rate'].replace(' ',''))
        early_amount = float(request.form['early_amount'].replace(' ',''))
        early_month = int(request.form['early_month'])
        early_year = int(request.form['early_year'])
        mode = request.form['mode']
        
        print(f"DEBUG: Расчет досрочного погашения - principal: {principal}, down_payment: {down_payment}, years: {years}, rate: {rate}")
        print(f"DEBUG: early_amount: {early_amount}, early_month: {early_month}, early_year: {early_year}, mode: {mode}")
        
        # Выполняем расчет досрочного погашения
        result, error = calculate_early_repayment(
            principal, down_payment, years, rate, 
            early_amount, early_month, early_year, mode
        )
        
        if error:
            print(f"DEBUG: Ошибка в calculate_early_repayment: {error}")
            return {'error': error}, 400
        
        # Получаем оригинальный график платежей
        original_schedule, _ = calculate_payment_schedule(principal, down_payment, years, rate)
        
        # Формируем ответ
        response_data = {
            'new_monthly_payment': format_number(result['new_monthly_payment']),
            'new_term': round(result['new_term'], 1),
            'interest_savings': format_number(result['interest_savings']),
            'term_reduction': round(result['term_reduction'], 1),
            'original_schedule': original_schedule,
            'new_schedule': result['new_schedule']
        }
        
        print(f"DEBUG: Успешный ответ: {response_data}")
        return response_data
        
    except Exception as e:
        print(f"DEBUG: Исключение в calculate_early_repayment_route: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'error': f'Ошибка расчета: {str(e)}'}, 400

# Запуск приложения в режиме отладки
if __name__ == '__main__':
    app.run(debug=True) 